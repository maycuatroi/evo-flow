import pandas as pd
from openpyxl import load_workbook

from evoflow.entities.data_manipulate.file_operator.file import File


class DataFrameFile(File):
    """
    Handle pandas DataFrame files
    """

    def save(self, file_path=None) -> bool:
        # pylint: disable=abstract-class-instantiated
        writer = pd.ExcelWriter(file_path, engine="xlsxwriter")

        # skip first row
        self.data.to_excel(
            writer, sheet_name="Sheet1", startrow=1, header=False, index=False
        )
        # pylint: disable=no-member
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]
        # Add a header format.
        header_format = workbook.add_format(
            {"bold": True, "fg_color": "#03a5fc", "border": 1}
        )

        for col_num, value in enumerate(self.data.columns.values):
            # write to second row
            worksheet.write(0, col_num, value, header_format)
            column_len = self.data[value].astype(str).str.len().max()
            column_len = max(column_len, len(value))
            worksheet.set_column(col_num, col_num, column_len)
        writer.save()
        return True

    # def get_info(self) -> str:
    #     wb = load_workbook(self.file_path)
    #     properties = wb.properties
    #     return self.data.head()

    def __init__(self, **args):
        super().__init__(**args)
        self.data: pd.DataFrame
        if self.file_path is not None:
            self.__meta_data__ = load_workbook(self.file_path).properties
        else:
            self.__meta_data__ = None
